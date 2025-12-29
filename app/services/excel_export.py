import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
from app.models.quotation import Quotation


def generate_quotation_excel(quotation: Quotation) -> bytes:
    """
    Generate Excel file for quotation.
    CONVERTED FROM: Django views.export_quotation_excel logic
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    
    # Header styling
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Quotation - {quotation.project_code}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = header_alignment
    
    # Add some space
    ws['A2'] = ""
    
    # Headers
    headers = ['Product Name', 'Unit Price (₹)', 'Quantity', 'Line Total (₹)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    row = 4
    total = Decimal('0.00')
    for line in quotation.lines:
        ws.cell(row=row, column=1, value=line.product_name)
        ws.cell(row=row, column=2, value=float(line.unit_price))
        ws.cell(row=row, column=3, value=line.qty)
        ws.cell(row=row, column=4, value=float(line.line_total))
        total += line.line_total
        row += 1
    
    # Add empty row
    row += 1
    
    # Total row
    ws.cell(row=row, column=3, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total)).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()
